# import_data

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from tqdm import tqdm

DB_CONFIG = {
    'dbname': 'bober',
    'user': 'postgres',
    'password': 'admin',
    'host': 'localhost',
    'port': '5432'
}


def parse_excel(file_path):
    """Анализирует Excel-файл и возвращает структурированные данные"""
    df = pd.read_excel(file_path)
    wb = load_workbook(file_path)  # Для доступа к гиперссылкам
    ws = wb.active
    # Очищаем данные - удаляем полностью пустые строки
    df = df.dropna(how='all').reset_index(drop=True)
    # Извлекаем сущности из второго столбца
    df['entity'] = df.iloc[:, 1].str.extract(r'^(.*?)(?:\s+тип\s+\d+)?$')[0]
    df['category'] = df.iloc[:, 14]  # Столбец 15 (индекс 14)
    df['subcategory'] = df.iloc[:, 11]  # Столбец 12 (индекс 11)
    data = {}
    current_entity = None
    current_object = None

    for index, row in df.iterrows():
        # Пропускаем заголовки (первые 4 строки)
        if index < 4:
            continue

        # Обновляем текущую сущность
        if pd.notna(row['entity']):
            current_entity = row['entity']
            if current_entity not in data:
                data[current_entity] = {
                    'category': row['category'],
                    'subcategory': row['subcategory'],
                    'characteristics': {},
                    'objects': {}
                }

        # Проверяем, что текущая сущность определена
        if current_entity is None:
            continue  # Пропускаем строки, если сущность не определена

        # Обновляем текущий объект
        object_name = row.iloc[1]
        if pd.notna(object_name):
            current_object = object_name
            if current_object not in data[current_entity]['objects']:
                data[current_entity]['objects'][current_object] = {}

        # Проверяем, что текущий объект определен
        if current_object is None:
            continue  # Пропускаем строки, если объект не определен

        # Сохраняем ОКПД2 как характеристику
        okpd2_value = row.iloc[2]
        if pd.notna(okpd2_value):
            data[current_entity]['objects'][current_object]['ОКПД2'] = okpd2_value

        # Сохраняем Коды КТРУ и ККН как характеристики
        ktru_cell = ws.cell(row=index + 2, column=13)  # 13-й столбец (KTRU)
        kkn_cell = ws.cell(row=index + 2, column=14)   # 14-й столбец (KKN)
        ktru_text = row.iloc[12] if pd.notna(row.iloc[12]) else None
        kkn_text = row.iloc[13] if pd.notna(row.iloc[13]) else None
        ktru_hyperlink = ktru_cell.hyperlink.target if ktru_cell.hyperlink else None
        kkn_hyperlink = kkn_cell.hyperlink.target if kkn_cell.hyperlink else None

        if ktru_text:
            data[current_entity]['objects'][current_object]['КТРУ текст'] = ktru_text
        if ktru_hyperlink:
            data[current_entity]['objects'][current_object]['КТРУ ссылка'] = ktru_hyperlink
        if kkn_text:
            data[current_entity]['objects'][current_object]['ККН текст'] = kkn_text
        if kkn_hyperlink:
            data[current_entity]['objects'][current_object]['ККН ссылка'] = kkn_hyperlink

        # Пропускаем строки без характеристик
        characteristic = row.iloc[5]
        if pd.isna(characteristic):
            continue

        # Обработка единиц измерения
        unit = row.iloc[10] if pd.notna(row.iloc[10]) and row.iloc[10] != '-' else None

        # Обработка значений (min, max, обычное значение)
        min_val = row.iloc[6] if pd.notna(row.iloc[6]) else None
        max_val = row.iloc[7] if pd.notna(row.iloc[7]) else None

        # Формируем значение
        if min_val is not None and max_val is not None:
            value = f"{min_val}...{max_val}"
        elif min_val is not None:
            value = f"≥{min_val}"
        elif max_val is not None:
            value = f"≤{max_val}"
        else:
            value = row.iloc[8] if pd.notna(row.iloc[8]) else row.iloc[9] if pd.notna(row.iloc[9]) else None

        # Добавляем характеристику если есть значение
        if value is not None:
            # Добавляем характеристику в список характеристик сущности
            if characteristic not in data[current_entity]['characteristics']:
                data[current_entity]['characteristics'][characteristic] = {
                    'unit': unit,
                    'data_type': row.iloc[4] if pd.notna(row.iloc[4]) else None
                }
            # Добавляем значение характеристики для объекта
            data[current_entity]['objects'][current_object][characteristic] = value

    return data


def init_database():
    """Инициализирует структуру базы данных"""
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = True
    cur = conn.cursor()
    try:
        # Удаляем существующие таблицы
        cur.execute("DROP TABLE IF EXISTS object_values")
        cur.execute("DROP TABLE IF EXISTS objects")
        cur.execute("DROP TABLE IF EXISTS entity_characteristics")
        cur.execute("DROP TABLE IF EXISTS entities")
        cur.execute("DROP TABLE IF EXISTS subcategories")
        cur.execute("DROP TABLE IF EXISTS categories")

        # Создаем таблицы
        cur.execute("""
            CREATE TABLE categories (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) UNIQUE NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE subcategories (
                id SERIAL PRIMARY KEY,
                category_id INTEGER NOT NULL REFERENCES categories(id) ON DELETE CASCADE,
                name text NOT NULL,
                UNIQUE(category_id, name)
            )
        """)
        cur.execute("""
            CREATE TABLE entities (
                id SERIAL PRIMARY KEY,
                name VARCHAR(255) UNIQUE NOT NULL,
                description TEXT,
                category VARCHAR(255),
                subcategory VARCHAR(255),
                subcategory_id INTEGER REFERENCES subcategories(id) ON DELETE SET NULL
            )
        """)
        cur.execute("""
            CREATE TABLE entity_characteristics (
                id SERIAL PRIMARY KEY,
                entity_id INTEGER NOT NULL REFERENCES entities(id) ON DELETE CASCADE,
                name text NOT NULL,
                data_type VARCHAR(50),
                unit VARCHAR(50),
                UNIQUE(entity_id, name)
            )
        """)
        cur.execute("""
            CREATE TABLE objects (
                id SERIAL PRIMARY KEY,
                entity_id INTEGER NOT NULL REFERENCES entities(id) ON DELETE CASCADE,
                name VARCHAR(255) NOT NULL,
                UNIQUE(entity_id, name)
            )
        """)
        cur.execute("""
            CREATE TABLE object_values (
                id SERIAL PRIMARY KEY,
                object_id INTEGER NOT NULL REFERENCES objects(id) ON DELETE CASCADE,
                characteristic_id INTEGER NOT NULL REFERENCES entity_characteristics(id),
                value TEXT,
                UNIQUE(object_id, characteristic_id)
            )
        """)
        print("База данных успешно инициализирована")
    finally:
        cur.close()
        conn.close()


def import_to_database(data):
    """Импортирует данные в БД"""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        # 1. Импорт категорий и подкатегорий
        categories = set()
        subcategories = set()
        for entity in data.values():
            if pd.notna(entity['category']):
                categories.add((entity['category'],))
            if pd.notna(entity['subcategory']):
                subcategories.add((entity['category'], entity['subcategory']))
        # Вставляем категории
        cur.executemany(
            "INSERT INTO categories (name) VALUES (%s) ON CONFLICT (name) DO NOTHING",
            categories
        )
        # Вставляем подкатегории
        for cat, subcat in subcategories:
            cur.execute("""
                INSERT INTO subcategories (category_id, name)
                SELECT id, %s FROM categories WHERE name = %s
                ON CONFLICT (category_id, name) DO NOTHING
            """, (subcat, cat))
        # 2. Импорт сущностей и их данных
        for entity_name, entity_data in tqdm(data.items(), desc="Импорт сущностей"):
            # Получаем ID подкатегории
            cur.execute("""
                SELECT sc.id FROM subcategories sc
                JOIN categories c ON sc.category_id = c.id
                WHERE c.name = %s AND sc.name = %s
            """, (entity_data['category'], entity_data['subcategory']))
            subcategory_id = cur.fetchone()[0] if cur.rowcount > 0 else None
            # Вставляем сущность
            cur.execute("""
                INSERT INTO entities (name, category, subcategory, subcategory_id)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (name) DO UPDATE 
                SET category = EXCLUDED.category,
                    subcategory = EXCLUDED.subcategory,
                    subcategory_id = EXCLUDED.subcategory_id
                RETURNING id
            """, (entity_name, entity_data['category'], entity_data['subcategory'], subcategory_id))
            entity_id = cur.fetchone()[0]
            # 3. Импорт характеристик сущности
            characteristics_map = {}  # Для хранения соответствия имен характеристик их ID
            for char_name, char_data in entity_data['characteristics'].items():
                cur.execute("""
                    INSERT INTO entity_characteristics (entity_id, name, data_type, unit)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (entity_id, name) DO UPDATE
                    SET data_type = EXCLUDED.data_type,
                        unit = EXCLUDED.unit
                    RETURNING id
                """, (entity_id, char_name, char_data.get('data_type'), char_data.get('unit')))
                char_id = cur.fetchone()[0]
                characteristics_map[char_name] = char_id
            # 4. Импорт объектов и их значений
            for obj_name, obj_data in entity_data['objects'].items():
                cur.execute("""
                    INSERT INTO objects (entity_id, name)
                    VALUES (%s, %s)
                    ON CONFLICT (entity_id, name) DO NOTHING
                    RETURNING id
                """, (entity_id, obj_name))
                obj_id = cur.fetchone()[0] if cur.rowcount > 0 else None
                if not obj_id:
                    cur.execute("""
                        SELECT id FROM objects 
                        WHERE entity_id = %s AND name = %s
                    """, (entity_id, obj_name))
                    obj_id = cur.fetchone()[0]
                # 5. Импорт значений характеристик
                for char_name, value in obj_data.items():
                    if char_name == 'ОКПД2':
                        characteristic_name = 'ОКПД2'
                    else:
                        characteristic_name = char_name
                    if characteristic_name not in characteristics_map:
                        cur.execute("""
                            INSERT INTO entity_characteristics (entity_id, name, data_type, unit)
                            VALUES (%s, %s, %s, %s)
                            ON CONFLICT (entity_id, name) DO UPDATE
                            SET data_type = EXCLUDED.data_type,
                                unit = EXCLUDED.unit
                            RETURNING id
                        """, (entity_id, characteristic_name, None, None))
                        char_id = cur.fetchone()[0]
                        characteristics_map[characteristic_name] = char_id
                    else:
                        char_id = characteristics_map[characteristic_name]
                    cur.execute("""
                        INSERT INTO object_values (object_id, characteristic_id, value)
                        VALUES (%s, %s, %s)
                        ON CONFLICT (object_id, characteristic_id)
                        DO UPDATE SET value = EXCLUDED.value
                    """, (obj_id, char_id, str(value)))
        conn.commit()
        print("\nИмпорт данных успешно завершен!")
        # Проверка количества записей
        cur.execute("SELECT COUNT(*) FROM entity_characteristics")
        print(f"Характеристик импортировано: {cur.fetchone()[0]}")
        cur.execute("SELECT COUNT(*) FROM object_values")
        print(f"Значений характеристик импортировано: {cur.fetchone()[0]}")
    except Exception as e:
        conn.rollback()
        print(f"\nОшибка при импорте: {e}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == '__main__':
    # 1. Инициализация БД
    print("Инициализация структуры БД...")
    init_database()
    # 2. Импорт данных из Excel
    file_path = 'doc.xlsx'  # Укажите путь к вашему файлу
    print(f"\nАнализ файла {file_path}...")
    data = parse_excel(file_path)
    # Вывод информации о данных перед импортом
    print("\nНайдены следующие данные:")
    for entity_name, entity_data in data.items():
        print(f"\nСущность: {entity_name}")
        print(f"Категория: {entity_data['category']}")
        print(f"Подкатегория: {entity_data['subcategory']}")
        print(f"Характеристики: {len(entity_data['characteristics'])}")
        print(f"Объекты: {len(entity_data['objects'])}")
        for obj_name, obj_data in entity_data['objects'].items():
            print(f"  Объект '{obj_name}': {len(obj_data)} значений характеристик")
    # 3. Импорт в БД
    print("\nНачинаем импорт в базу данных...")
    import_to_database(data)